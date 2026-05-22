import os
import re
import unicodedata
import logging
from datetime import datetime
import openpyxl
from pypdf import PdfReader, PdfWriter

os.makedirs("logs", exist_ok=True)
logging.basicConfig(
    filename="logs/procesos.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

VALOR_ACTIVACION = {
    "chk_tramite_a": "/X", "chk_tramite_b": "/X", "chk_tramite_c": "/Yes_xuhu",
    "chk_tramite_d": "/Yes_xuhu", "chk_tramite_e": "/Yes_xuhu", "chk_tramite_f": "/Yes_xuhu",
    "chk_objeto_inicial": "/Yes_xuhu", "chk_objeto_prorroga": "/Yes_xuhu",
    "chk_objeto_modificacion": "/Yes_xuhu", "chk_objeto_revalidacion": "/Yes_xuhu",
    "chk_urb_a": "/Yes_xuhu", "chk_urb_b": "/Yes_xuhu", "chk_urb_c": "/Yes_xuhu",
    "chk_subd_a": "/Yes_xuhu", "chk_subd_b": "/Yes_xuhu", "chk_subd_c": "/Yes_xuhu",
    "chk_const_a": "/Yes_xuhu", "chk_const_b": "/Yes_xuhu", "chk_const_c": "/Yes_xuhu",
    "chk_const_d": "/Yes_xuhu", "chk_const_e": "/Yes_xuhu", "chk_const_f": "/Yes_xuhu",
    "chk_const_g_total": "/Yes_xuhu", "chk_const_g_parcial": "/Yes_xuhu",
    "chk_const_h": "/Yes_xuhu", "chk_const_i": "/Yes_xuhu",
    "chk_uso_a": "/Yes_xuhu", "chk_uso_b": "/Yes_xuhu", "chk_uso_c": "/Yes_xuhu",
    "chk_uso_d": "/Yes_xuhu", "chk_uso_e": "/Yes_xuhu",
    "chk_area_a": "/Yes_xuhu", "chk_area_b": "/Yes_xuhu", "chk_area_c": "/Yes_xuhu",
    "chk_vis_a": "/Yes_xuhu", "chk_vis_b": "/Yes_xuhu", "chk_vis_c": "/Yes_xuhu",
    "chk_cultural_a": "/Yes_xuhu", "chk_cultural_b": "/Yes_xuhu",
    "chk_sost_a": "/Yes_xuhu", "chk_sost_b": "/Yes_xuhu", "chk_sost_c": "/Yes_xuhu",
    "chk_clima_e": "/Yes_xuhu", "chk_clima_f": "/Yes_xuhu",
    "chk_suelo_a": "/Yes_xuhu", "chk_suelo_b": "/Yes_xuhu", "chk_suelo_c": "/Yes_xuhu",
    "chk_plani_a": "/Yes_xuhu", "chk_plani_b": "/Yes_xuhu", "chk_plani_c": "/Yes_xuhu",
}

def quitar_tildes(texto):
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')

def normalizar(valor):
    if valor is None: return ""
    valor = str(valor).strip()
    if valor.endswith(".0"): valor = valor[:-2]
    return re.sub(r"\s+", " ", quitar_tildes(valor)).upper()

def buscar_archivo_por_inicio(carpeta, inicio):
    if not os.path.exists(carpeta): return None
    for archivo in os.listdir(carpeta):
        if archivo.upper().startswith(inicio.upper()) and archivo.lower().endswith(".xlsx"):
            return os.path.join(carpeta, archivo)
    return None

def buscar_registro_en_todas_las_hojas(ruta_excel, criterio_busqueda):
    datos = {}
    if not ruta_excel or not os.path.exists(ruta_excel): return datos
    criterio = normalizar(criterio_busqueda)
    
    wb = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)
    for hoja_nombre in wb.sheetnames:
        hoja = wb[hoja_nombre]
        primera_fila = next(hoja.iter_rows(max_row=1), None)
        if not primera_fila: continue
        
        encabezados = [normalizar(celda.value).lower() for celda in primera_fila]
        idx_cedula, idx_codigo = -1, -1
        for i, cab in enumerate(encabezados):
            if any(p in cab for p in ["cedula", "documento", "nit"]): idx_cedula = i
            if any(p in cab for p in ["codigo", "predial", "catastral"]): idx_codigo = i

        vacias = 0
        for fila_celdas in hoja.iter_rows(min_row=2):
            fila = [c.value for c in fila_celdas]
            if not any(v is not None for v in fila):
                vacias += 1
                if vacias >= 3: break
                continue
            vacias = 0

            val_ced = normalizar(fila[idx_cedula]) if idx_cedula != -1 and idx_cedula < len(fila) else ""
            val_cod = normalizar(fila[idx_codigo]) if idx_codigo != -1 and idx_codigo < len(fila) else ""

            if criterio == val_ced or criterio == val_cod:
                datos.update({cab: fila[i] if i < len(fila) else None for i, cab in enumerate(encabezados) if cab})
                break
    wb.close()
    return datos

def llenar_pdf(ruta_plantilla, ruta_salida, campos_texto, casillas, radios):
    reader = PdfReader(ruta_plantilla)
    writer = PdfWriter(clone_from=reader)
    
    writer.update_page_form_field_values(writer.pages[0], campos_texto, auto_regenerate=False)
    
    valores_marca = {campo: VALOR_ACTIVACION.get(campo) for campo in casillas if VALOR_ACTIVACION.get(campo)}
    if valores_marca:
        writer.update_page_form_field_values(writer.pages[0], valores_marca, auto_regenerate=False)
        
    if radios:
        writer.update_page_form_field_values(writer.pages[0], radios, auto_regenerate=False)
        
    writer.set_need_appearances_writer(True)
    os.makedirs(os.path.dirname(ruta_salida) or ".", exist_ok=True)
    with open(ruta_salida, "wb") as f:
        writer.write(f)

def obtener_datos_predio(criterio: str):
    ruta_catastro = os.path.join("excels", "CB1_REGISTRO CATASTRO BCA.xlsx")
    catastro = buscar_registro_en_todas_las_hojas(ruta_catastro, criterio)
    if not catastro:
        return None

    tipo_suelo = "URBANO"
    for k, v in catastro.items():
        if any(p in str(k).lower() for p in ["suelo", "clase", "clasificacion", "clasificación"]):
            if "RURAL" in normalizar(v): tipo_suelo = "RURAL"

    valores_concat = " ".join(normalizar(v) for v in catastro.values() if v)
    if "VEREDA" in valores_concat: tipo_suelo = "RURAL"

    secundaria = {}
    ruta_secundaria = buscar_archivo_por_inicio("excels", "VEREDAS" if tipo_suelo == "RURAL" else "BARRIOS")
    if ruta_secundaria:
        secundaria = buscar_registro_en_todas_las_hojas(ruta_secundaria, criterio)

    datos_finales = {**catastro, **secundaria}
    
    direccion, barrio, vereda, matricula, codigo, nombre = "", "", "", "", "", ""
    for k, v in datos_finales.items():
        clave = str(k).lower()
        if "direccion" in clave: direccion = normalizar(v)
        elif "barrio" in clave: barrio = normalizar(v)
        elif "vereda" in clave: vereda = normalizar(v)
        elif "matricula" in clave: matricula = normalizar(v)
        elif "codigo" in clave or "predial" in clave: codigo = normalizar(v)
        elif any(p in clave for p in ["propietario", "nombre", "interesado"]): nombre = normalizar(v)

    return {
        "nombre": nombre,
        "direccion": direccion,
        "barrio_vereda": vereda if tipo_suelo == "RURAL" else barrio,
        "matricula": matricula,
        "codigo_predial": codigo,
        "tipo_suelo": tipo_suelo
    }

def ejecutar_llenado_pdf(criterio: str, datos_usuario: dict):
    if not datos_usuario:
        raise ValueError("datos_usuario está vacío o es None")

    ruta_pdf = os.path.join("plantillas_pdf", "formulario_base.pdf")
    logging.info(f"Procesando PDF definitivo para el criterio: {criterio}")

    # ⚠️ CORRECCIÓN AQUÍ (claves correctas)
    campos_texto = {
        "txt_direccion": datos_usuario.get("direccion", ""),
        "txt_barrio": datos_usuario.get("barrio_vereda", "") if datos_usuario.get("tipo_suelo") != "RURAL" else "",
        "txt_vereda": datos_usuario.get("barrio_vereda", "") if datos_usuario.get("tipo_suelo") == "RURAL" else "",
        "txt_matricula": datos_usuario.get("matricula", ""),
        "txt_codigo_predial": datos_usuario.get("codigo_predial", ""),
    }

    casillas = []

    # ✔️ Suelo automático
    if datos_usuario.get("tipo_suelo") == "RURAL":
        casillas.append("chk_suelo_b")
    else:
        casillas.append("chk_suelo_a")

    # ✔️ Casillas extra (seguro)
    if datos_usuario.get("casillas_extra"):
        casillas.extend(datos_usuario["casillas_extra"])

    # ✔️ Radios
    radios = {}
    if datos_usuario.get("grupo_clima"):
        radios["grupo_clima"] = datos_usuario["grupo_clima"]

    nombre_archivo = f"formulario_{criterio}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    ruta_salida = os.path.join("generados", nombre_archivo)

    llenar_pdf(ruta_pdf, ruta_salida, campos_texto, casillas, radios)

    logging.info(f"PDF generado correctamente en: {ruta_salida}")

    return ruta_salida, nombre_archivo
